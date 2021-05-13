import dash  # version 1.13.1       # stacked bar çizdiriyor. Axis isimleri eklenecek.
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output, ALL, State, MATCH, ALLSMALLER
import plotly.express as px

from openpyxl import Workbook,load_workbook
import numpy as np
workbook = load_workbook(r"C:\Users\Hamza\Desktop\GAMS OUTPUTS\Case2_Abdurrahman.xlsx")
sheet = workbook.active
sheet = workbook["güç dengesi"]
list_home1=[]
columns=["Time","Pbuy_t","P_ev_used","P_pv_used","P_bat_used","P_ev_ch","P_bat_ch","P_load_t"]
for value in sheet.iter_rows(min_row=87,
   max_row=96,
   min_col=2,
   max_col=97,
   values_only=True):
   list_home1.append(value)


from pandas import DataFrame
df_home1 = DataFrame(list_home1)

df_home1=df_home1.T

df_home1.drop([6,1],axis=1,inplace=True)

df_home1.columns=columns

df_home1          #buraya kadar df çıkarıldı.

app = dash.Dash(__name__)
app.layout = html.Div([
    html.H1("Select every component for each house respectively", style={'text-align': 'center'}),  #app4'den ekledim.

    html.Div(children=[
        html.Button('Add Chart', id='add-chart', n_clicks=0),   #chart ekleme butonu
    ]),
    html.Div(id='container', children=[])                      #grafik için konteynır
])


@app.callback(
    Output('container', 'children'),
    [Input('add-chart', 'n_clicks')],   #Tıklama sayısına(n'e) göre tetiklenir
    [State('container', 'children')]    #?
)
def display_graphs(n_clicks, div_children):
    new_child = html.Div(
        style={'width': '45%', 'display': 'inline-block', 'outline': 'thin lightgrey solid', 'padding': 10},
        children=[
            dcc.Graph(
                id={
                    'type': 'dynamic-graph',    #dinamik grafik türü
                    'index': n_clicks
                },
                figure={}
            ),
            dcc.RadioItems(  #radiobuttons   #grafik tipi seçimi
                id={
                    'type': 'dynamic-choice',
                    'index': n_clicks
                },
                options=[{'label': 'Bar Chart', 'value': 'bar'},
                         {'label': 'Line Chart', 'value': 'line'},
                         {'label': 'Pie Chart', 'value': 'pie'}],
                value='bar',  #default seçtik
            ),
            dcc.Dropdown(
                id={
                    'type': 'dynamic-dpn-s',
                    'index': n_clicks
                },
                options=[{'label': s, 'value': s} for s in np.sort(df_home1.columns.unique())],
                multi=True,
                value=["Pbuy_t", "P_ev_used"],   #default seçtik.
            )
            # dcc.Dropdown(
            #     id={
            #         'type': 'dynamic-dpn-num',
            #         'index': n_clicks
            #     },
            #     options=[{'label': n, 'value': n} for n in ['detenues', 'under trial', 'convicts', 'others']],
            #     value='convicts',
            #     clearable=False
            # )
            # dcc.Dropdown(
            #     id={
            #         'type': 'dynamic-dpn-ctg',
            #         'index': n_clicks
            #     },
            #     options=[{'label': c, 'value': c} for c in df_home1["Time"]],
            #     value=None,
            #     clearable=False
            # )

        ]
    )
    div_children.append(new_child)
    return div_children


@app.callback(
    Output({'type': 'dynamic-graph', 'index': MATCH}, 'figure'),
    [Input(component_id={'type': 'dynamic-dpn-s', 'index': MATCH}, component_property='value'),  #for match, there is all and allsmaller
     #Input(component_id={'type': 'dynamic-dpn-num', 'index': MATCH}, component_property='value'),
     Input({'type': 'dynamic-choice', 'index': MATCH}, 'value')]
)
def update_graph(s_value, chart_choice):  #respectively above(in inputs)
    print(s_value)
    #dff = df[df['state'].isin(s_value)]

    if chart_choice == 'bar':
        #dff = dff.groupby([ctg_value], as_index=False)[['detenues', 'under trial', 'convicts', 'others']].sum()
        import plotly.graph_objects as go
        a = df_home1.loc[:, "Time"]
        b = df_home1.loc[:, "Pbuy_t"]
        c = df_home1.loc[:, "P_ev_used"]
        d = df_home1.loc[:, "P_bat_used"]
        e = df_home1.loc[:, "P_ev_ch"]
        f = df_home1.loc[:, "P_bat_ch"]
        h = df_home1.loc[:, "P_pv_used"]
        g = df_home1.loc[:, "P_load_t"]

        fig = go.Figure(data=[         #home1
            go.Bar(name='Pbuy_t', x=a, y=b),   #burada y için bir şekilde s_value'yu çiz demem lazım. Grafik tipiyle oynanabilir.
            go.Bar(name='P_ev_used', x=a, y=c),
            go.Bar(name='P_bat_used', x=a, y=d),
            go.Bar(name='P_ev_ch', x=a, y=e),
            go.Bar(name='P_bat_ch', x=a, y=f),
            go.Bar(name='P_load_t', x=a, y=g),
            go.Bar(name='P_pv_used', x=a, y=h)])

        # Change the bar mode
        fig.update_layout(barmode='overlay', template='gridon')
        #fig.show()
        return fig
    elif chart_choice == 'line':
        if len(s_value) == 0:
            return {}
        else:
            dff = dff.groupby([ctg_value, 'year'], as_index=False)
            [['detenues', 'under trial', 'convicts', 'others']].sum()
            fig = px.line(dff, x='year', y=num_value, color=ctg_value)
            return fig
    elif chart_choice == 'pie':
        fig = px.pie(dff, names=ctg_value, values=num_value)
        return fig


if __name__ == '__main__':
    app.run_server(debug=True,port=8052,use_reloader=False)



# https://youtu.be/4gDwKYaA6ww